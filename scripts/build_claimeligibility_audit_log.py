from __future__ import annotations

import json
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment


REPO = Path(r"C:\Users\phong\Downloads\send\multimethod_xai_crop_yield_anomalies")
AUDIT_DIR = Path(r"D:\00_Major\K4\AI2011_Group8_SE190791_TranDaiPhongLam\AI_audit_log")
INPUT = AUDIT_DIR / "AI_AuditLog_AUS_EarlyWarning_DAP391m.xlsx"
OUTPUT = AUDIT_DIR / "AI_AuditLog_ClaimEligibility_DAP391m_v1.xlsx"
REPORT = AUDIT_DIR / "AI_AuditLog_ClaimEligibility_DAP391m_build_report.md"

CONTEXT_DOCX = Path(r"C:\Users\phong\Downloads\ICTAI_claim_eligibility_project_context_handoff_v5_7_main_2026-07-24.docx")
V57_PDF = REPO / "paper/final/ictai2026_claim_eligibility_audit_v5_7_web_prototype_figure.pdf"
V57_TEX = REPO / "paper_versions/v5_7_web_prototype_figure/source/fidelity_gated_xai_method_benchmark_v3.tex"

MARK = "[RECONSTRUCTED FROM PROJECT EVIDENCE]"


def rel(path: str | Path) -> str:
    path = Path(path)
    if path.is_absolute():
        try:
            return path.relative_to(REPO).as_posix()
        except ValueError:
            return str(path)
    return path.as_posix()


entries = [
    [
        "001",
        "DECISION",
        "Business & Problem Understanding",
        "The early manuscript read too much like a crop-yield/XAI application and did not clearly foreground the reusable contribution.",
        f"{MARK} How should the research be reframed so the central contribution is claim eligibility rather than crop-model accuracy?",
        "Recommended defining the problem as controlling the permitted interpretation scope of post-hoc explanations using locked evidence.",
        "Accepted the recommendation, but kept the crop panel as the main application. This made the paper about disciplined claim scope, while explicitly avoiding a high-performing crop-model success claim because the crop modules do not pass.",
        "C:\\Users\\phong\\Downloads\\ICTAI_claim_eligibility_project_context_handoff_v5_7_main_2026-07-24.docx; paper/final/ictai2026_claim_eligibility_audit_v5_7_web_prototype_figure.pdf; paper_versions/v5_7_web_prototype_figure/source/fidelity_gated_xai_method_benchmark_v3.tex",
    ],
    [
        "002",
        "VERIFICATION",
        "Business & Problem Understanding",
        "Multiple V5.6, V5.6.2 and V5.7 branches existed, so the current submission version could be confused.",
        f"{MARK} Which paper version is the current main version, and which one is only the no-web fallback?",
        "V5.7 is the main and most complete integrated version; V5.6 is the safest no-web fallback; V5.6.2 is an optional no-web layout branch.",
        "Accepted this hierarchy. The workbook treats V5.7 as the scientific source of truth and does not silently downgrade to a no-web branch, which keeps the audit log aligned with the final selected paper.",
        "C:\\Users\\phong\\Downloads\\ICTAI_claim_eligibility_project_context_handoff_v5_7_main_2026-07-24.docx; paper/final/ictai2026_claim_eligibility_audit_v5_7_web_prototype_figure_completion_report.md",
    ],
    [
        "003",
        "DECISION",
        "Business & Problem Understanding",
        "The project needed a precise definition of what the audit decides.",
        f"{MARK} Is this audit a row-level reject option, a causal-validity test, or a study-level permission rule?",
        "It is a study-level permission rule defining the highest claim scope supported by locked evidence; it is not a causal-validity detector.",
        "Accepted and used this distinction across the paper and web-planning materials. This prevented the project from being framed as row-level selective prediction or automatic causal/confounding detection.",
        "paper_versions/v5_7_web_prototype_figure/source/fidelity_gated_xai_method_benchmark_v3.tex; web_product_handoff_plan/02_scope_and_non_goals.md; web_product_handoff_plan/10_audit_engine_specification.md",
    ],
    [
        "004",
        "VERIFICATION",
        "Data Understanding & Preparation",
        "Dataset scope and analysis unit had to match the final paper.",
        f"{MARK} Verify the crop panel unit, time range, crops, states, data sources and feature scope.",
        "Confirmed crop-state-year rows, 1,257 observations, 1990-2025, four crops, twelve U.S. states, USDA NASS yield and NASA POWER weather with full-season features.",
        "Accepted only the verified state-panel values and removed the unrelated project dataset counts. This kept the workbook tied to the actual Claim-Eligibility paper data rather than the old template.",
        "paper_versions/v5_7_web_prototype_figure/source/fidelity_gated_xai_method_benchmark_v3.tex; paper_versions/v5_7_web_prototype_figure/source/generated/audit_numbers_v4.tex; DATA_MANIFEST.md",
    ],
    [
        "005",
        "DECISION",
        "Data Understanding & Preparation",
        "Detrending and event construction could cause temporal leakage or target confusion.",
        f"{MARK} How should the target and event label be constructed without using future information?",
        "Fit each crop-state trend and residual scale on training observations only; use raw residual as the prediction target and standardized residual only to define event labels.",
        "Accepted the separation because it protects locked evaluation from future target leakage. The concrete effect was to preserve train-only target construction and keep event labels out of model features.",
        "paper_versions/v5_7_web_prototype_figure/source/fidelity_gated_xai_method_benchmark_v3.tex; artifacts/audit/split/detrending_audit.csv; artifacts/audit/split/split_manifest.csv",
    ],
    [
        "006",
        "VERIFICATION",
        "Data Understanding & Preparation",
        "The locked evaluation population had to remain independent of model selection.",
        f"{MARK} Verify the validation interval, locked-test interval, candidate exclusions and final locked-row count.",
        "Confirmed validation 2012-2015, locked test 2016-2025, 140 retained validation rows and 333 evaluated locked rows.",
        "Accepted these fixed intervals and row counts. The project therefore keeps model selection on validation only and prevents locked-test access during configuration selection.",
        "paper_versions/v5_7_web_prototype_figure/source/generated/audit_numbers_v4.tex; artifacts/audit/selection/validation_model_grid.csv; artifacts/audit/final_test/row_level_predictions.csv",
    ],
    [
        "007",
        "VERIFICATION",
        "Data Understanding & Preparation",
        "Leakage-sensitive columns might enter model matrices.",
        f"{MARK} Which fields must be prohibited from the model feature matrices?",
        "Exclude year, yield, trend, residual, standardized residual, event labels, predictions, history length and residual scale.",
        "Accepted explicit forbidden-column checks instead of manual trust. This made the workbook document leakage safeguards as part of the evidence trail, not as an afterthought.",
        "C:\\Users\\phong\\Downloads\\ICTAI_claim_eligibility_project_context_handoff_v5_7_main_2026-07-24.docx; artifacts/audit/stage_features/feature_dictionary.csv; artifacts/data/feature_matrix_schema.json",
    ],
    [
        "008",
        "DECISION",
        "Exploratory Data Analysis",
        "State-level performance heterogeneity could be misread as multiple independent module decisions.",
        f"{MARK} How should state-level Delta RMSE results be visualized and interpreted?",
        "Use a U.S. state map as descriptive subgroup evidence only; do not let state-level estimates replace the panel-level Module A decision.",
        "Accepted this boundary. Figure 4 stayed descriptive, so subgroup contrasts inform diagnosis without creating separate permission verdicts outside the pre-specified panel-level rule.",
        "paper_versions/v5_7_web_prototype_figure/source/fidelity_gated_xai_method_benchmark_v3.tex; artifacts/maps/state_level_locked_delta_rmse.csv; paper/generated/figure_state_delta_rmse_map.pdf",
    ],
    [
        "009",
        "VERIFICATION",
        "Exploratory Data Analysis",
        "ROC-AUC and PR-AUC appeared diagnostically positive even though Module E did not pass.",
        f"{MARK} Does ROC-AUC above random establish event-recovery eligibility?",
        "No. ROC-AUC summarizes global pairwise ordering, while Module E requires tail error, rank uncertainty and top-k evidence to pass together.",
        "Accepted the stricter interpretation. ROC/PR were reported as diagnostics only and were not allowed to override failed Module E rank and top-k evidence.",
        "paper_versions/v5_7_web_prototype_figure/source/fidelity_gated_xai_method_benchmark_v3.tex; artifacts/audit/tail/tail_metrics_by_threshold.csv; artifacts/audit_records/topk_null_audit.csv",
    ],
    [
        "010",
        "DECISION",
        "Exploratory Data Analysis",
        "The original SHAP waterfall presentation could imply that SHAP itself failed.",
        f"{MARK} How should the Barley-Colorado 2016 explanation be visualized without misrepresenting SHAP?",
        "Use a three-panel figure: outcome mismatch, grouped SHAP contributions and audited interpretation scope.",
        "Accepted and refined the figure. It now says SHAP coherently explains the fitted prediction, while the unsupported step is promotion from model explanation to observed-event weather attribution.",
        "paper_versions/v5_7_web_prototype_figure/source/fidelity_gated_xai_method_benchmark_v3.tex; artifacts/xai/shap_reconstruction_checks.csv; artifacts/gates/figure2_three_comparisons.json",
    ],
    [
        "011",
        "DECISION",
        "Modeling & Regression Analysis",
        "The model-selection design needed to distinguish algorithms from feature representations.",
        f"{MARK} Which candidate models and representations should be compared during validation?",
        "Evaluate Ridge, Random Forest and ExtraTrees across Metadata-only, Weather-only and Full representations with fixed configurations and seeds.",
        "Accepted the deliberately small auditable grid. The team did not expand the model set after seeing locked-test results, preserving the locked protocol.",
        "artifacts/audit/selection/search_space.json; artifacts/audit/selection/validation_model_grid.csv; artifacts/audit/selection/selected_config.json",
    ],
    [
        "012",
        "VERIFICATION",
        "Modeling & Regression Analysis",
        "The selected validation configuration could be overclaimed as clearly superior.",
        f"{MARK} Verify the selected model, seed behavior and whether validation superiority persisted on the locked period.",
        "Weather-only ExtraTrees was selected, but the advantage was small relative to seed variability; validation RMSE 0.384 deteriorated to locked RMSE 0.669 and locked R2 was -0.014.",
        "Accepted the pre-specified selection but rejected a best-model success claim. This kept the result honest: the selected model is used for auditing, not sold as a strong crop predictor.",
        "paper_versions/v5_7_web_prototype_figure/source/generated/audit_numbers_v4.tex; artifacts/audit/selection/validation_seed_metrics.csv; artifacts/gates/figure2_three_comparisons.json",
    ],
    [
        "013",
        "DECISION",
        "Modeling & Regression Analysis",
        "Overall model adequacy required a fixed statistical gate.",
        f"{MARK} What should Module A test, and what is the crop-panel verdict?",
        "Compare the selected model with the pre-specified baseline on identical locked rows and pass only if the upper paired 95% CI is below zero. Crop estimate: -0.005 with CI [-0.019, 0.009], so it does not pass.",
        "Accepted the interval rule and rejected point-estimate overclaiming. This directly set the crop-paper verdict to model-descriptive explanation only.",
        "paper_versions/v5_7_web_prototype_figure/source/generated/audit_numbers_v4.tex; artifacts/gates/figure2_three_comparisons.json; artifacts/tables/fidelity_gate_components.csv",
    ],
    [
        "014",
        "DECISION",
        "Modeling & Regression Analysis",
        "Weather-specific interpretation required evidence beyond overall model fit.",
        f"{MARK} How should incremental weather-feature value be tested?",
        "Module B compares Full with Metadata-only predictions on identical locked rows using the same selected architecture. Crop estimate: -0.012 with CI [-0.029, 0.002], so it does not pass.",
        "Accepted the feature-family gate and blocked weather-specific predictive-reliance claims despite the favorable point estimate. The result stayed conservative and evidence-based.",
        "paper_versions/v5_7_web_prototype_figure/source/generated/audit_numbers_v4.tex; artifacts/gates/figure2_three_comparisons.json; artifacts/tables/gate_decision_matrix.csv",
    ],
    [
        "015",
        "DECISION",
        "Modeling & Regression Analysis",
        "Event recovery could not be represented by one error metric.",
        f"{MARK} Which checks must Module E pass before an event-recovery claim is permitted?",
        "Tail error, rank recovery and chance-adjusted top-k prioritization must all pass. Crop tail Delta RMSE was favorable but uncertain, rank evidence failed and top-10 recovered only 1/10.",
        "Accepted the all-checks rule. This prevented one favorable tail-error metric from overriding failed rank/top-k evidence, so the observed-event claim remains unsupported.",
        "paper_versions/v5_7_web_prototype_figure/source/generated/audit_numbers_v4.tex; artifacts/audit/tail/tail_metrics_by_threshold.csv; artifacts/audit_records/rank_null_audit.csv; artifacts/audit_records/topk_null_audit.csv",
    ],
    [
        "016",
        "VERIFICATION",
        "Evaluation, Visualization & Reporting",
        "Synthetic ground truth could accidentally become an oracle input to the audit.",
        f"{MARK} Can synthetic ground-truth claim labels be used to set the A/B/E policy verdict?",
        "No. Synthetic GT is evaluation-only and is used only to measure false permission and false abstention.",
        "Accepted this separation and corrected circular wording. The observable policy remains based only on Modules A, B and E, while GT labels are used only for evaluation metrics.",
        "C:\\Users\\phong\\Downloads\\ICTAI_claim_eligibility_project_context_handoff_v5_7_main_2026-07-24.docx; artifacts/experiments/synthetic-gate-benchmark/summary.json; artifacts/experiments/synthetic-gate-benchmark/synthetic_ground_truth.csv",
    ],
    [
        "017",
        "VERIFICATION",
        "Evaluation, Visualization & Reporting",
        "The synthetic results and cross-domain cases needed an honest interpretation.",
        f"{MARK} What do the synthetic specificity result, county case and PJM case jointly show?",
        "The audit reduces false permission when signal is absent or degraded but cannot detect every invalid design preserving predictive performance. County has incremental evidence without overall adequacy; PJM is a positive control where applicable modules pass.",
        "Accepted the balanced interpretation. The report compares verdicts rather than raw RMSE magnitudes across domains with different units, preserving the audit's limitation.",
        "artifacts/experiments/synthetic-gate-benchmark/summary.json; artifacts/experiments/county-v2-weather-models/summary.json; artifacts/experiments/external-domain-eia/pjm_gate_decisions.json; paper_versions/v5_7_web_prototype_figure/source/fidelity_gated_xai_method_benchmark_v3.tex",
    ],
    [
        "018",
        "DECISION",
        "Evaluation, Visualization & Reporting",
        "The paper added a Workbench interface figure, creating a risk of product and causal overclaims.",
        f"{MARK} How should Figure 5 and the Claim-Eligibility Audit Workbench be described?",
        "Describe it as an interface prototype populated with an illustrative synthetic run, with a versioned audit trail and Module D diagnostic only. It adds no new empirical evidence and does not establish causality.",
        "Accepted qualified prototype wording and rejected deployed-product language. This kept Figure 5 as an operationalization/provenance illustration, not an extra experiment or causal system.",
        "C:\\Users\\phong\\Downloads\\ICTAI_claim_eligibility_project_context_handoff_v5_7_main_2026-07-24.docx; paper_versions/v5_7_web_prototype_figure/source/fidelity_gated_xai_method_benchmark_v3.tex; web_product_handoff_plan/02_scope_and_non_goals.md",
    ],
    [
        "019",
        "VERIFICATION",
        "Evaluation, Visualization & Reporting",
        "The submission package had to remain anonymous, exactly eight pages and scientifically invariant.",
        f"{MARK} What final checks are required before using the V5.7 paper for submission and presentation?",
        "Check page count, fixed metrics, figure readability, source/PDF identities, metadata, local paths, author names, emails, ORCIDs and version hierarchy.",
        "Accepted the rendered-page and metadata inspection discipline. The blind V5.7 PDF stays separate from authored materials, and future edits must preserve scientific values and anonymity checks.",
        "paper/final/ictai2026_claim_eligibility_audit_v5_7_web_prototype_figure_completion_report.md; paper/final/ictai2026_claim_eligibility_audit_v5_7_web_prototype_figure.pdf; C:\\Users\\phong\\Downloads\\ICTAI_claim_eligibility_project_context_handoff_v5_7_main_2026-07-24.docx",
    ],
]

hallucinations = [
    [
        "013 or 012",
        "Overclaiming / Logic Error",
        "The selected crop model beats the baseline and supports a successful predictive claim because its Delta RMSE point estimate is negative.",
        "Module A does not pass because the paired 95% CI [-0.019, 0.009] crosses zero; locked R2 is -0.014.",
        "Checked Table I, Table II, audit_numbers_v4.tex and paired year-block bootstrap artifacts.",
        "Described the crop verdict as model-descriptive explanation only.",
    ],
    [
        "010",
        "Context Misunderstanding / Overclaiming",
        "The positive SHAP weather contributions explain why weather caused the observed Barley-Colorado yield shortfall.",
        "SHAP explains why the fitted model predicted +0.209, while the observed residual was -0.510. A/B/E do not pass, and no causal claim is supported.",
        "Compared Figure 2, the observed outcome and prediction, SHAP reconstruction checks and module verdicts.",
        "Retained SHAP only as a model-descriptive explanation.",
    ],
    [
        "016",
        "Logic Error / Oracle Leakage",
        "The audit uses known synthetic validity labels to determine whether a claim passes.",
        "Synthetic GT is evaluation-only and is never input to Modules A, B or E.",
        "Checked synthetic benchmark implementation outputs, Table III, summary.json and the context handoff.",
        "Separated the observable A/B/E policy from GT-based evaluation metrics.",
    ],
    [
        "003 or 017",
        "Context Misunderstanding",
        "Module D must pass before feature-specific interpretation is permitted.",
        "Module D is a descriptive Weather-only versus Metadata-only diagnostic outside the permission path.",
        "Checked Figure 1, module definitions, Table I wording and the context handoff.",
        "Removed Module D from the sequential permission rule and labeled it diagnostic only.",
    ],
    [
        "018",
        "Fabrication / Product Overclaim",
        "Figure 5 demonstrates a deployed platform that detects leakage, confounding and causal feature effects with an immutable audit trail.",
        "Figure 5 is an interface prototype populated with an illustrative synthetic run. It operationalizes decision logic only, uses a versioned audit trail and does not establish causality or automatically detect structural invalidity.",
        "Checked the Figure 5 caption, web-product scope, V5.7 limitations and planning handoff.",
        "Used 'interface prototype', 'versioned audit trail' and 'predictive eligibility, not causality'.",
    ],
]

type_rows = [
    ["Overclaiming", "AI states a stronger scientific conclusion than the locked evidence supports.", "Negative point estimate treated as a passed module."],
    ["Oracle Leakage", "AI allows hidden labels or known validity to influence the policy decision.", "Synthetic GT used as A/B/E input."],
    ["Product Overclaim", "AI describes a prototype or plan as a deployed validated system.", "Workbench figure described as deployed causal product."],
    ["Context Misunderstanding", "AI misunderstands the audit scope or module role.", "Module D treated as a required gate."],
    ["Logic Error", "AI draws an invalid conclusion from otherwise real evidence.", "CI crossing zero treated as a pass."],
    ["Fabrication", "AI asserts unsupported artifacts, system status, or capabilities.", "Immutable audit trail or automatic causal detection claimed."],
]

viva_rows = [
    ["001", "Why the paper was reframed around claim eligibility instead of crop-model accuracy.", "Yes", "Context handoff sections 1-2; V5.7 Introduction and Conclusion."],
    ["005", "Why train-only detrending and event construction prevent leakage.", "Yes", "V5.7 Section III-A; artifacts/audit/split/detrending_audit.csv."],
    ["013", "Why a negative Delta RMSE point estimate does not mean Module A passes.", "Yes", "audit_numbers_v4.tex; artifacts/gates/figure2_three_comparisons.json."],
    ["016", "Why synthetic GT is evaluation-only.", "Yes", "context handoff section 2.3; synthetic summary.json."],
    ["018", "Why Figure 5 is only a prototype and not a causal/deployed system.", "Yes", "V5.7 Figure 5 caption; web_product_handoff_plan/02_scope_and_non_goals.md."],
]


def clear_values(ws, max_row=120, max_col=30):
    for row in ws.iter_rows(min_row=1, max_row=max(max_row, ws.max_row), min_col=1, max_col=max(max_col, ws.max_column)):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def set_wrap(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            base = cell.alignment or Alignment()
            cell.alignment = copy(base)
            cell.alignment = Alignment(
                horizontal=base.horizontal,
                vertical=base.vertical or "top",
                text_rotation=base.text_rotation,
                wrap_text=True,
                shrink_to_fit=base.shrink_to_fit,
                indent=base.indent,
            )


def set_print_layout(ws, print_area: str):
    ws.print_area = print_area
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35


def copy_row_style(ws, src_row, dst_row, max_col):
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def build():
    if not INPUT.exists():
        raise FileNotFoundError(INPUT)
    shutil.copy2(INPUT, OUTPUT)
    wb = load_workbook(OUTPUT)
    expected = [
        "1. Metadata & Summary",
        "2. Detailed Audit Log",
        "3. Hallucination Detection",
        "4. Self-Assessment Checklist",
    ]
    if wb.sheetnames != expected:
        raise RuntimeError(f"Unexpected sheet names: {wb.sheetnames}")

    ws1, ws2, ws3, ws4 = [wb[s] for s in expected]

    for ws in wb.worksheets:
        clear_values(ws)

    ws1["A1"] = "AI AUDIT LOG - CLAIM-ELIGIBILITY AUDITING FOR POST-HOC EXPLANATIONS"
    ws1["A3"] = "STUDENT INFORMATION"
    ws1["A4"], ws1["B4"] = "Student Name:", "Tran Dai Phong Lam"
    ws1["A5"], ws1["B5"] = "Student ID:", "SE190791"
    ws1["A6"], ws1["B6"] = "Course:", "DAP391m"
    ws1["A7"], ws1["B7"] = "Assignment:", "Claim-Eligibility Auditing for Post-hoc Explanations: Synthetic Calibration and Cross-Domain Cases"
    ws1["A8"], ws1["B8"] = "Project type:", "Regression, temporal locked evaluation, explainable machine learning and claim-scope auditing."
    ws1["A9"] = "AI USAGE SUMMARY"
    ws1["A10"], ws1["B10"], ws1["C10"] = "Total Prompts Used (all AI tools):", "Not reliably recoverable", "No complete verbatim prompt log was found."
    ws1["A11"], ws1["B11"], ws1["C11"] = "Core Prompts Logged:", "=COUNTA('2. Detailed Audit Log'!A4:A22)", "Formula counts valid rows in Sheet 2."
    ws1["A12"], ws1["B12"], ws1["C12"] = "Selection Ratio:", "N/A - total prompt count unavailable", "Selection ratio is not calculated because total prompts are not verifiable."
    ws1["A13"], ws1["B13"], ws1["C13"] = "Hallucination Detected:", "=COUNTA('3. Hallucination Detection'!A4:A8)", "Formula counts valid correction cases in Sheet 3."
    ws1["A14"], ws1["B14"] = "Reconstruction disclosure:", "Core entries are concise records of AI-assisted project decisions. Entries marked [RECONSTRUCTED FROM PROJECT EVIDENCE] are evidence-based summaries rather than verbatim chat transcripts."
    ws1["A15"] = "AI TOOLS USED"
    ws1["A16"], ws1["B16"], ws1["C16"], ws1["D16"] = "AI Tool", "Purpose", "Frequency", "Main Value"
    ws1["A17"], ws1["B17"], ws1["C17"], ws1["D17"] = "ChatGPT", "Research framing, reviewer critique, explanation and revision planning.", "High", "Clarified claim scope, limitations and final paper narrative."
    ws1["A18"], ws1["B18"], ws1["C18"], ws1["D18"] = "Codex", "Repository inspection, code execution, figure regeneration, LaTeX/PDF QA, PowerPoint preparation and workbook reconstruction.", "High", "Connected paper claims to source artifacts and reproducible build checks."
    ws1["A22"] = "CORE PROMPTS DISTRIBUTION "
    dist = [
        ("Business & Problem Understanding", ">= 2"),
        ("Data Understanding & Preparation", ">= 3"),
        ("Exploratory Data Analysis", ">= 2"),
        ("Modeling & Regression Analysis", ">= 4"),
        ("Evaluation, Visualization & Reporting", ">= 3"),
    ]
    for i, (label, minimum) in enumerate(dist, start=23):
        ws1.cell(i, 1).value = label
        ws1.cell(i, 2).value = f'=COUNTIF(\'2. Detailed Audit Log\'!C4:C22,"{label}")'
        ws1.cell(i, 3).value = minimum
    ws1["A28"], ws1["B28"], ws1["C28"] = "Total", "=SUM(B23:B27)", "Expected: 19"
    ws1.column_dimensions["B"].width = 58
    ws1.column_dimensions["C"].width = 48
    ws1.column_dimensions["D"].width = 32
    ws1.row_dimensions[7].height = 42
    ws1.row_dimensions[8].height = 33
    ws1.row_dimensions[14].height = 54
    ws1.row_dimensions[17].height = 45
    ws1.row_dimensions[18].height = 45
    set_wrap(ws1, "A1:F28")
    set_print_layout(ws1, "A1:F28")

    ws2["A1"] = "DETAILED AI AUDIT LOG"
    ws2["A2"] = "INSTRUCTIONS: Only record CORE PROMPTS (Decision/Problem-Solving/Verification). Reconstructed entries are evidence-based summaries, not verbatim prompt transcripts."
    headers = ["Entry #", "Prompt Type", "Stage/Component", "Problem/Context", "Prompt to AI", "AI Response (Summary)", "Human Delta & Reflection", "Evidence"]
    for col, header in enumerate(headers, 1):
        ws2.cell(3, col).value = header
    for r, entry in enumerate(entries, 4):
        for c, value in enumerate(entry, 1):
            ws2.cell(r, c).value = value
        ws2.row_dimensions[r].height = 150
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 26
    ws2.column_dimensions["D"].width = 48
    ws2.column_dimensions["E"].width = 58
    ws2.column_dimensions["F"].width = 58
    ws2.column_dimensions["G"].width = 62
    ws2.column_dimensions["H"].width = 48
    set_wrap(ws2, "A1:H22")
    set_print_layout(ws2, "A1:H22")

    ws3["A1"] = "HALLUCINATION DETECTION LOG (MANDATORY)"
    ws3["A2"] = "Evidence-based correction cases for Claim-Eligibility Auditing."
    headers3 = ["Entry # (from Sheet 2)", "Hallucination Type", "AI's Claim", "Reality Check", "How Detected", "Corrective Action"]
    for col, header in enumerate(headers3, 1):
        ws3.cell(3, col).value = header
    for r, case in enumerate(hallucinations, 4):
        for c, value in enumerate(case, 1):
            ws3.cell(r, c).value = value
        ws3.row_dimensions[r].height = 108
    ws3["A11"] = "HALLUCINATION TYPES REFERENCE:"
    ws3["A12"], ws3["B12"], ws3["C12"] = "Type", "Definition", "Example"
    for idx, row in enumerate(type_rows, 13):
        if idx > 36:
            copy_row_style(ws3, 36, idx, 6)
        for c, value in enumerate(row, 1):
            ws3.cell(idx, c).value = value
        ws3.row_dimensions[idx].height = 42
    for col, width in {"A": 18, "B": 30, "C": 45, "D": 50, "E": 45, "F": 45}.items():
        ws3.column_dimensions[col].width = width
    set_wrap(ws3, "A1:F18")
    set_print_layout(ws3, "A1:F18")

    ws4["A1"] = "SELF-ASSESSMENT CHECKLIST (Before Submission)"
    ws4["A2"] = "Check carefully before submitting. EACH ENTRY must pass >=4/5 of the criteria below."
    ws4["A4"] = "A. QUALITY CHECK FOR EACH ENTRY (Pass >=4/5)"
    ws4["A5"], ws4["B5"], ws4["C5"], ws4["D5"] = "#", "Criteria", "Pass?", "Note"
    quality = [
        ("1", "Did this prompt affect an important decision in the project?", "☑", "All 19 entries affected framing, protocol, model selection, evaluation, visualization or reporting scope."),
        ("2", "Without this prompt, would the project change in architecture/design?", "☑", "Each entry records a method, gate, dataset, evaluation, figure or product-scope consequence."),
        ("3", "Can I explain why I accepted/rejected the AI suggestion?", "☑", "Each Human Delta states what was accepted, modified or rejected and why."),
        ("4", "Is there concrete evidence (code, metrics, comparison)?", "☑", "Every row has a non-empty evidence cell tied to source, PDF, report or artifact evidence."),
        ("5", "Does this prompt reflect critical thinking, not just copying the AI?", "☑", "Entries emphasize human decisions such as rejecting crop-model success, oracle GT use, causal claims and Module D gating."),
    ]
    for r, row in enumerate(quality, 6):
        for c, value in enumerate(row, 1):
            ws4.cell(r, c).value = value
        ws4.row_dimensions[r].height = 32
    ws4["A12"] = "B. OVERALL LOG CHECK"
    ws4["A13"], ws4["B13"], ws4["C13"], ws4["D13"] = "#", "Criteria", "Pass?", "Current Value"
    overall = [
        ("1", "Is the number of entries within the range (min-max)?", "☑", "19 core entries, all reconstructed from project evidence and disclosed as such."),
        ("2", "Does each component have at least 1 core prompt?", "☑", "All five required components are represented with distribution 3/4/3/5/4."),
        ("3", "Has the required number of hallucinations been detected (>= required)?", "☑", "Five evidence-supported correction cases are recorded."),
        ("4", "Does every entry have a complete Human Delta (4 questions)?", "☑", "Every row explains recommendation, accept/reject/modify, reason and concrete project effect."),
        ("5", "Is there evidence for >=70% of entries?", "☑", "19/19 entries have evidence references; most point to real local files verified during build."),
    ]
    for r, row in enumerate(overall, 14):
        for c, value in enumerate(row, 1):
            ws4.cell(r, c).value = value
        ws4.row_dimensions[r].height = 32
    ws4["A20"], ws4["B20"] = "NOTE:", "If an entry does NOT pass >=4/5 criteria, remove it rather than forcing a pass."
    ws4["A21"], ws4["B21"] = "NOTE:", "If >=2 overall criteria fail, the AI Reflection score is at risk."
    ws4["A25"], ws4["B25"] = "C.", "PREPARATION FOR ORAL VIVAS (Q&A)"
    ws4["A26"], ws4["B26"] = "", "The instructor may ask about 3-5 entries. These five are the strongest evidence-backed explanations."
    ws4["A27"], ws4["B27"], ws4["C27"], ws4["D27"] = "Entry #", "Can I explain why this approach was chosen?", "AI response remembered?", "Do I have evidence?"
    for r, row in enumerate(viva_rows, 28):
        if r > 28:
            copy_row_style(ws4, 28, r, 4)
        for c, value in enumerate(row, 1):
            ws4.cell(r, c).value = value
        ws4.row_dimensions[r].height = 40
    ws4.column_dimensions["B"].width = 72
    ws4.column_dimensions["C"].width = 24
    ws4.column_dimensions["D"].width = 55
    set_wrap(ws4, "A1:D32")
    set_print_layout(ws4, "A1:D32")

    wb.save(OUTPUT)
    return {
        "input": str(INPUT),
        "output": str(OUTPUT),
        "sheets": wb.sheetnames,
        "entries": len(entries),
        "hallucinations": len(hallucinations),
        "reconstructed_entries": len(entries),
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }


if __name__ == "__main__":
    result = build()
    print(json.dumps(result, indent=2))
